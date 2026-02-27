import openpyxl
from openpyxl.utils import get_column_letter

class ExcelService:
    
    @staticmethod
    def read_supporting_excel(file_path):
        """Read Decathlon SKU, quantity, and cost data from Excel using header names"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Find column indices by header name
            header_map = {}
            # Expected headers and their internal keys
            column_defs = {
                'decathlon_sku': ['Decathlon SKU', 'Decathlon SKU #', 'SKU #', 'SKU', 'Model Code', 'Item Code'],
                'model': ['Model', 'Model Code', 'Item Code', 'Model #', 'Item #', 'Article #', 'Style', 'Article Code', 'Product Code'],
                'item_description': ['Item Description', 'Description', 'Product Description', 'Product Name', 'Name', 'Title'],
                'quantity': ['QTY', 'Qty', 'Quantity', 'Units'],
                'unit_cost': ['Unit Cost without VAT', 'Foreign FOB', 'Unit Cost', 'Cost', 'Cost Price'],
                'unit_retail': ['Unit Retail With VAT', 'Unit Retail', 'Retail Price', 'RRP', 'Unit Price'],
                'barcode': ['Barcode', 'EAN', 'UPC', 'GTIN', 'International Code']
            }
            
            # Find header row by searching first 15 rows
            header_row_idx = None
            print(f"[DEBUG] Searching for headers in Excel file...")
            for r_idx in range(1, 16):
                try:
                    row_values = list(ws.iter_rows(min_row=r_idx, max_row=r_idx, values_only=True))[0]
                    if not row_values: continue
                    
                    print(f"[DEBUG] Row {r_idx} values: {row_values[:10]}")  # Show first 10 columns
                    
                    found_keys = {}
                    for idx, cell_value in enumerate(row_values):
                        if cell_value is None: continue
                        cell_str = str(cell_value).strip().lower()
                        # Use partial matching for longer headers, exact for shorter ones
                        for key, possibilities in column_defs.items():
                            if key in found_keys: continue
                            for p in possibilities:
                                p_low = p.lower()
                                if cell_str == p_low or (len(p_low) > 5 and p_low in cell_str):
                                    found_keys[key] = idx
                                    print(f"[DEBUG] Found '{key}' at column {idx} (header: '{cell_value}')")
                                    break
                    
                    # If we found at least Decathlon SKU or Barcode + QTY/Cost, it's likely the header row
                    if 'decathlon_sku' in found_keys or ('barcode' in found_keys and len(found_keys) >= 2):
                        header_map = found_keys
                        header_row_idx = r_idx
                        print(f"[DEBUG] ✓ Header row found at index {r_idx}: {header_map}")
                        break
                except:
                    continue
            
            if header_row_idx is None:
                 print(f"[DEBUG] Header search failed. Defaulting to row 1.")
                 header_row_idx = 1
                 # Last ditch effort on row 1
                 headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                 for idx, cell_value in enumerate(headers):
                    if not cell_value: continue
                    cell_str = str(cell_value).strip().lower()
                    for key, possibilities in column_defs.items():
                        if key in header_map: continue
                        if any(p.lower() == cell_str for p in possibilities):
                            header_map[key] = idx

            data = []
            row_count = 0
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                # Stop if row is completely empty
                if not any(row): continue
                
                row_count += 1
                # Helper to safely get value by key
                def get_val(key):
                    if key in header_map and header_map[key] < len(row):
                        val = row[header_map[key]]
                        return val if val is not None else ''
                    return ''

                # Extract and normalize
                d_sku_val = get_val('decathlon_sku')
                d_sku = str(d_sku_val).strip() if d_sku_val is not None else ''
                if d_sku.lower().endswith('.0'): d_sku = d_sku[:-2]
                
                model_val = get_val('model')
                model = str(model_val).strip() if model_val is not None else ''
                if model.lower().endswith('.0'): model = model[:-2]
                
                item_desc_val = get_val('item_description')
                item_description = str(item_desc_val).strip() if item_desc_val is not None else ''
                
                barcode_val = get_val('barcode')
                barcode = ''
                if barcode_val:
                    try:
                        if isinstance(barcode_val, (int, float)):
                            barcode = "{:.0f}".format(barcode_val)
                        else:
                            barcode = str(barcode_val).strip()
                            if barcode.lower().endswith('.0'):
                                barcode = barcode[:-2]
                    except:
                        barcode = str(barcode_val).strip()

                if d_sku or model or barcode:
                    qty_val = get_val('quantity')
                    try:
                        qty = float(qty_val) if qty_val is not None and qty_val != '' else 0
                    except: qty = 0
                        
                    cost_val = get_val('unit_cost')
                    try:
                        cost = float(cost_val) if cost_val is not None and cost_val != '' else 0.0
                    except: cost = 0.0
                        
                    retail_val = get_val('unit_retail')
                    try:
                        retail = float(retail_val) if retail_val is not None and retail_val != '' else 0.0
                    except: retail = 0.0

                    item_data = {
                        'decathlon_sku': d_sku,
                        'model': model,
                        'item_description': item_description,
                        'barcode': barcode,
                        'quantity': qty,
                        'unit_cost': cost,
                        'unit_retail': retail,
                        'color_size': f"000|{d_sku}" if d_sku else f"000|{model}"
                    }
                    
                    # Log first few items for debugging
                    if len(data) < 3:
                        print(f"[DEBUG] Excel row {row_count}: Decathlon SKU='{d_sku}', Model='{model}', Barcode='{barcode}'")
                    
                    data.append(item_data)
            
            print(f"[DEBUG] Successfully read {len(data)} items from Excel (processed {row_count} rows)")
            return data
        except Exception as e:
            print(f"[ERROR] read_supporting_excel failed: {str(e)}")
            raise Exception(f"Excel reading error: {str(e)}")
    
    @staticmethod
    def generate_erp_excel(invoice_data, invoice_items, supplier_name, business_unit_code):
        """Generate ERP-ready Excel file with PO Creation and IM Creation sheets"""
        try:
            wb = openpyxl.Workbook()
            
            # ===== SHEET 1: PO CREATION (existing template) =====
            ws_po = wb.active
            ws_po.title = "PO Creation"
            
            # Define headers for PO Creation
            headers_po = [
                'Company', 'Brand', 'MCU', 'InvoiceNumber', 'Albaran', 'BOX#',
                'DateYYYYMMDD', 'Itemcode', 'Color|Size', 'Barcode', 'QTY',
                'Local FOB', 'Foreign Cur', 'Foreign FOB', 'Unit Retail'
            ]
            
            ws_po.append(headers_po)
            
            # Add data rows for PO Creation
            for item in invoice_items:
                row = [
                    '06002',              # Company
                    '54',                 # Brand
                    business_unit_code,   # MCU
                    invoice_data.get('invoice_number', ''), # InvoiceNumber
                    '',                   # Albaran
                    '',                   # BOX#
                    invoice_data.get('invoice_date', ''),   # DateYYYYMMDD
                    item.get('itemcode', ''),               # Itemcode
                    item.get('color_size', ''),             # Color|Size
                    item.get('barcode', ''),                # Barcode
                    item.get('quantity', 0),                # QTY
                    '',                                     # Local FOB
                    invoice_data.get('currency', 'QAR'),    # Foreign Cur
                    item.get('unit_cost', 0),               # Foreign FOB (The unit Cost)
                    item.get('unit_retail', 0)              # Unit Retail
                ]
                ws_po.append(row)
            
            # Auto-adjust column widths for PO sheet
            for column in ws_po.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_po.column_dimensions[column_letter].width = adjusted_width
            
            # ===== SHEET 2: IM CREATION =====
            ws_im = wb.create_sheet("IM Creation")
            
            # Define headers for IM Creation
            headers_im = [
                'Itemcode', 'Desc. Line 1', 'Desc. Line 2', 'mancode', 'brand', 
                'season', 'supplier', 'section', 'family', 'subfamily', 
                'feature code', 'alternate code', 'HS Code', 'COO'
            ]
            
            ws_im.append(headers_im)
            
            # Add data rows for IM Creation
            for item in invoice_items:
                # Split description if it's longer than 30 characters
                full_desc = item.get('item_description', '')
                if len(full_desc) > 30:
                    desc_line_1 = full_desc[:30]
                    desc_line_2 = full_desc[30:]
                else:
                    desc_line_1 = full_desc
                    desc_line_2 = ''
                
                row = [
                    item.get('itemcode', ''),           # Itemcode
                    desc_line_1,                        # Desc. Line 1
                    desc_line_2,                        # Desc. Line 2
                    item.get('mancode', ''),            # mancode
                    item.get('brand_code', ''),         # brand
                    item.get('season', ''),             # season
                    item.get('supplier_code', ''),      # supplier
                    item.get('section', ''),            # section
                    item.get('family', ''),             # family
                    item.get('subfamily', ''),          # subfamily
                    '',                                 # feature code (empty)
                    item.get('alternate_code', ''),     # alternate code
                    '',                                 # HS Code (empty)
                    ''                                  # COO (empty)
                ]
                ws_im.append(row)
            
            # Auto-adjust column widths for IM sheet
            for column in ws_im.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_im.column_dimensions[column_letter].width = adjusted_width
            
            return wb
        except Exception as e:
            raise Exception(f"Excel generation error: {str(e)}")
