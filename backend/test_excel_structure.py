import openpyxl
import os

# Find latest Excel file
files = [f for f in os.listdir('uploads') if f.endswith('.xlsx')]
files.sort(key=lambda x: os.path.getmtime(os.path.join('uploads', x)), reverse=True)

if not files:
    print("No Excel files found")
    exit()

latest = files[0]
print(f"Examining: {latest}")
print("=" * 60)

wb = openpyxl.load_workbook(f'uploads/{latest}', data_only=True)
ws = wb.active

# Show first 5 rows
for i in range(1, 6):
    row_values = [cell.value for cell in ws[i]]
    print(f"Row {i}: {row_values[:15]}")  # First 15 columns

print("=" * 60)

# Try to find Decathlon SKU column
print("\nSearching for 'Decathlon SKU' or similar...")
for row_idx in range(1, 10):
    row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
    for col_idx, cell_value in enumerate(row):
        if cell_value and 'decathlon' in str(cell_value).lower():
            print(f"Found at Row {row_idx}, Column {col_idx}: '{cell_value}'")
        if cell_value and 'sku' in str(cell_value).lower():
            print(f"Found SKU-related at Row {row_idx}, Column {col_idx}: '{cell_value}'")
